#!/usr/bin/env python3
"""Build the Hormuz-2026 reserve-intelligence workbook with LIVE Excel formulas.
Run: python3 build/build_xlsx.py  -> OUTPUT_hormuz2026_reserve_intelligence.xlsx
Data as-of 26 Jun 2026 (workflow run wf_9642a98b-45e). Figures are wire-attributed
where primary IEA/EIA hosts were proxy-blocked — re-fetch before external use.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette (Bloomberg/GS terminal) -------------------------------------
INK   = "0B1F2A"; PANEL = "112C3A"; ACCENT = "F5A623"; AMBER = "FFB020"
GOOD  = "1F7A4D"; WARN = "B07A1E"; BAD = "9E2B25"; GREY = "8AA0AD"; WHITE = "FFFFFF"
H_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
T_FONT = Font(name="Calibri", color="1B2B34", size=10)
hdr_fill  = PatternFill("solid", fgColor=INK)
sub_fill  = PatternFill("solid", fgColor=PANEL)
thin = Side(style="thin", color="D5DDE2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

CONF_FILL = {
    "Verified-official":      PatternFill("solid", fgColor="D6F0E0"),
    "Inferred-triangulated":  PatternFill("solid", fgColor="FBEED2"),
    "Modelled-EST":           PatternFill("solid", fgColor="F6D9C7"),
    "Absent":                 PatternFill("solid", fgColor="E6EAED"),
}

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H_FONT; cell.fill = hdr_fill; cell.alignment = CEN; cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ============================================================ NOTES =========
ws = wb.active; ws.title = "READ ME"
ws["A1"] = "Hormuz 2026 — Crude Reserve, Redistribution & Restocking Intelligence"
ws["A1"].font = Font(bold=True, size=14, color="0B1F2A")
notes = [
    "",
    "As-of: 26 Jun 2026  |  War start 28 Feb 2026  |  Window T = 118 days  |  Brent $74.43  |  Hormuz contested-but-flowing ~4.8 mb/d",
    "",
    "THESIS VERDICT: the 'global reserves are conserved, just redistributed producer-side' hypothesis is REFUTED (strong form).",
    "  rho = producer build / consumer draw = 120 / 615 = 0.195  (band 0.15-0.26; an order of magnitude below conservation = 1.0).",
    "  Mechanism: Gulf tankage saturated in ~1-2 weeks, so blocked output was SHUT IN at the wellhead (~700-1,080 mb), not stockpiled.",
    "",
    "LIVE FORMULAS: every computed cell is a real Excel formula, not a pasted number. Edit a 'days' value and months/deltas/at-risk recompute.",
    "  Months   = days / 30.44",
    "  Delta days = Today - Before ;  Delta % = (Today - Before) / Before",
    "  Eff. at-risk cover = Today days x (1 - Hormuz-dependence)   [worst-case zero-replacement BOUND, not a coverage estimate]",
    "  rho (ledger) = Producer build / Consumer draw   (live on the Ledger sheet)",
    "",
    "CONFIDENCE LADDER:  Verified-official > Inferred-triangulated > Modelled-EST > Absent('-').  A model is at most ONE source class.",
    "",
    "CAVEAT (load-bearing): primary IEA OMR / EIA weekly hosts were egress-proxy-403-blocked at run time, so many cells are",
    "WIRE-ATTRIBUTED (one tier below Verified). Re-pull IEA/EIA directly before putting these decimals in front of a manager.",
    "",
    "Sheets:  Consumer | Producer | Key Indicators | Ledger | Throughput | Trajectory | Scenarios | Footnotes",
]
for i, t in enumerate(notes, 3):
    ws.cell(row=i, column=1, value=t).font = Font(size=10, color=("9E2B25" if "REFUTED" in t or "CAVEAT" in t else "1B2B34"),
                                                  bold=("THESIS" in t or "LIVE FORMULAS" in t or "CAVEAT" in t))
autosize(ws, [140])

# ============================================================ CONSUMER ======
ws = wb.create_sheet("Consumer")
cols = ["Country","Before (days)","Before (months)","Today (days)","Today (months)",
        "Δ days","Δ %","Hormuz-dep %","Eff. at-risk cover","Confidence","Source & Date"]
ws.append(cols); style_header(ws, 1, len(cols))
# country, before_days, today_days, hormuz_dep(0-1), confidence, source
consumer = [
    ("USA",          None, None, 0.08, "Absent",                "EIA WPSR / IEA OMR May-26; net-exporter, metric undefined"),
    ("Germany",      None, 99.1, 0.05, "Modelled-EST",          "BMWE/EBV + IEA, 11 Mar 26 (before=110 imputed)"),
    ("France",       98.0, 98.0, 0.11, "Inferred-triangulated", "IEA/Euronews/Eurostat, Apr 26"),
    ("UK",          104.0, 90.0, 0.07, "Inferred-triangulated", "GOV.UK/IEA/Al Jazeera, Mar 26"),
    ("Italy",        None, 90.0, 0.18, "Inferred-triangulated", "IEA Italy Oil Security / EU Dir 2009/119"),
    ("Spain",       120.0, 91.0, 0.06, "Modelled-EST",          "CORES/IEA/Al Jazeera, Mar 26"),
    ("Netherlands",  None, None, 0.06, "Absent",                "IEA/COVA; net-exporter, metric undefined"),
    ("Poland",       None, 86.5, 0.13, "Modelled-EST",          "IEA Poland Oil Security + model"),
    ("Japan",       200.0,205.0, 0.90, "Inferred-triangulated", "S&P/METI-ANRE, 9 May 26"),
    ("South Korea", 208.0,205.0, 0.90, "Inferred-triangulated", "IEA tool/CSIS/Sedaily, May 26"),
    ("China",       122.5,130.0, 0.46, "Inferred-triangulated", "Reuters-Kemp/EIA, 2026"),
    ("India",        74.0, 76.0, 0.30, "Inferred-triangulated", "Min. Puri 76-80d, 8 Jun 26"),
    ("Thailand",     None,117.0, 0.55, "Inferred-triangulated", "Thai Energy Ministry, 12 May 26 (hard on-ground ~56d)"),
    ("Philippines",  None,46.47, 0.95, "Inferred-triangulated", "PH DOE OIMB weekly, 12 Jun 26"),
    ("Indonesia",    None, 21.0, 0.20, "Inferred-triangulated", "Jakarta Globe/ESDM, Jun 26 (capacity-capped ~25d)"),
    ("Vietnam",      None, 40.3, 0.88, "Inferred-triangulated", "VietnamNet D-basis, May 26 (model ~16d; divergence preserved)"),
]
for i, (c, b, t, dep, conf, src) in enumerate(consumer, start=2):
    r = i
    ws.cell(r,1,c).alignment = LEFT
    ws.cell(r,2, b if b is not None else "-").alignment = CEN
    ws.cell(r,3, (f"=B{r}/30.44" if b is not None else "-")).alignment = CEN
    ws.cell(r,4, t if t is not None else "-").alignment = CEN
    ws.cell(r,5, (f"=D{r}/30.44" if t is not None else "-")).alignment = CEN
    ws.cell(r,6, (f"=D{r}-B{r}" if (b is not None and t is not None) else "-")).alignment = CEN
    ws.cell(r,7, (f"=(D{r}-B{r})/B{r}" if (b is not None and t is not None) else "-")).alignment = CEN
    ws.cell(r,7).number_format = "0.0%"
    ws.cell(r,8, dep).alignment = CEN; ws.cell(r,8).number_format = "0%"
    ws.cell(r,9, (f"=D{r}*(1-H{r})" if t is not None else "-")).alignment = CEN
    cc = ws.cell(r,10, conf); cc.alignment = CEN; cc.fill = CONF_FILL[conf]
    ws.cell(r,11, src).alignment = LEFT
    for cidx in range(1,12):
        ws.cell(r,cidx).border = BORDER;
        if cidx in (3,5): ws.cell(r,cidx).number_format = "0.0"
ws.freeze_panes = "A2"
autosize(ws, [13,12,13,12,13,9,8,11,15,20,46])

# ============================================================ PRODUCER ======
ws = wb.create_sheet("Producer")
cols = ["Country","Export-cover (days) deployable","Export-cover (days) total","Usable in-Hormuz (mb)",
        "Usable out-Hormuz (mb)","Usable abroad (mb)","Storage-util %","Bypass-adj export (mb/d)",
        "Shut-in (mb/d)","Confidence","Source & Date"]
ws.append(cols); style_header(ws, 1, len(cols))
producer = [
    ("Saudi Arabia", 7.7, 31.9, 150, 30, 18, 0.92, 4.5, 2.3, "Inferred-triangulated","Argus/Bloomberg/Kayrros, Mar-Jun 26"),
    ("UAE",          9.0, 14.0,   0, 42,  6, 0.90, 1.8, 1.4, "Inferred-triangulated","IEA OMR via The National, 24 Jun 26"),
    ("Iran",         6.0, 57.0, 106,  0,  0, 0.91, 0.1,1.45, "Inferred-triangulated","Kpler/Al Jazeera, Jun 26"),
    ("Iraq",         0.3,  4.7,  17,  1,  0, 0.95, 0.3, 1.5, "Inferred-triangulated","Bloomberg/IEA OMR Apr/Basra Oil Co"),
    ("Kuwait",       4.2, 16.0,  21,  0,7.5, 0.90, 0.0, 2.0, "Inferred-triangulated","Bloomberg/Zawya/The National, Apr-Jun 26"),
    ("Qatar",        0.0, 10.5,  10,  0,  0, 0.92, 0.0,0.95, "Modelled-EST",          "EIA Qatar brief + OilPrice (downgraded)"),
]
for i, row in enumerate(producer, start=2):
    c,dep,tot,inh,out,abr,util,byp,shut,conf,src = row
    vals = [c,dep,tot,inh,out,abr,util,byp,shut,conf,src]
    for cidx, v in enumerate(vals, 1):
        cell = ws.cell(i, cidx, v)
        cell.alignment = LEFT if cidx in (1,11) else CEN
        cell.border = BORDER
        if cidx == 7: cell.number_format = "0%"
        if cidx == 10: cell.fill = CONF_FILL[conf]
ws.freeze_panes = "A2"
autosize(ws, [14,16,15,15,15,13,11,13,12,20,40])

# ============================================================ KEY IND =======
ws = wb.create_sheet("Key Indicators")
cols = ["Indicator","Prewar / Before","Peak","Now (26 Jun)","Note"]
ws.append(cols); style_header(ws,1,len(cols))
ind = [
    ("Brent ($/bbl)","$72 (27 Feb); $65 late-25 anchor","~$120","$74.43","EIA RBRTE + Trading Economics"),
    ("US SPR (mb)","411 (31 Dec 25)","-","331 (19 Jun)","46.4% of ~714 nameplate; lowest since 1983"),
    ("IEA collective release (mb)","-","-","400 (US share 172)","decided 11 Mar; ~186 executed by 19 Jun"),
    ("VLCC ($/day MEG-China)","~$55k","~$423-470k","~$179,600","still >3x prewar"),
    ("War-risk hull premium","0.10-0.25%","2.5-5%","0.3-0.5%","eased post-ceasefire; reversible"),
    ("Hormuz flow (mb/d)","~15","-","4.8","contested-but-flowing; 32% of prewar"),
]
for i,row in enumerate(ind, start=2):
    for cidx,v in enumerate(row,1):
        cell=ws.cell(i,cidx,v); cell.alignment = LEFT; cell.border=BORDER
ws.freeze_panes="A2"; autosize(ws,[28,28,14,18,40])

# ============================================================ LEDGER ========
ws = wb.create_sheet("Ledger")
ws["A1"]="GLOBAL REDISTRIBUTION LEDGER — thesis test (all mb, window T=118d)"
ws["A1"].font=Font(bold=True,size=12,color="0B1F2A")
ws.append([])
ws.append(["Leg","mb","Note"]); style_header(ws,3,3)
legs = [
    ("Consumer / strategic stock draw", 615, "incl ~80 SPR + ~106 other-IEA strategic releases"),
    ("Demand destruction",              472, "2Q26 deliveries -5 mb/d y/y (IEA OMR Jun)"),
    ("Atlantic substitution (in)",      410, "rerouted East-of-Suez"),
    ("Producer build (stranded)",       120, "floating ~100 + onshore ~20"),
    ("Foregone output (shut-in)",       775, "time-weighted ~700-850 (flat-rate point ~1,080)"),
]
r=4
for name,mb,note in legs:
    ws.cell(r,1,name).alignment=LEFT; ws.cell(r,2,mb).alignment=CEN; ws.cell(r,3,note).alignment=LEFT
    for c in range(1,4): ws.cell(r,c).border=BORDER
    r+=1
# live rho + reconciliation
ws.cell(r+1,1,"rho = Producer build / Consumer draw").font=Font(bold=True)
ws.cell(r+1,2,f"=B7/B4").font=Font(bold=True,color="9E2B25"); ws.cell(r+1,2).number_format="0.000"
ws.cell(r+1,3,"REFUTED if << 1. (conservation threshold = 1.0)")
ws.cell(r+2,1,"War supply gap (mb)"); ws.cell(r+2,2,1200)
ws.cell(r+3,1,"Canonical observed net draw (mb)"); ws.cell(r+3,2,495)
ws.cell(r+4,1,"Identity check: DD + observed draw + Atlantic").font=Font(bold=True)
ws.cell(r+4,2,f"=B5+B{r+3}+B6").font=Font(bold=True)  # demand destruction + observed draw + atlantic
ws.cell(r+5,1,"Reconciliation residual (gap - identity)")
ws.cell(r+5,2,f"=B{r+2}-B{r+4}"); ws.cell(r+5,3,"~ -177 mb (~15%) = the 615-vs-495 error bar")
autosize(ws,[42,12,52])

# ============================================================ THROUGHPUT ====
ws = wb.create_sheet("Throughput")
cols=["Quantity","mb/d","Note"]; ws.append(cols); style_header(ws,1,3)
tp=[("Hormuz normal crude exit",15,"prewar baseline"),
    ("Bypass ceiling",6.5,"port-gated, not pipeline-gated"),
    ("  Yanbu / Petroline (Red Sea)",4.0,"~70% of all bypass; Yanbu port-limited"),
    ("  Fujairah / ADCOP (Gulf of Oman)",1.5,"true bypass; terminal at strait mouth"),
    ("  Kirkuk-Ceyhan (Med)",0.2,"north-only; near-zero true relief"),
    ("  Goreh-Jask",0.0,"idle since 2024"),
    ("Hormuz actual flow (now)",4.8,"contested-but-flowing"),
    ("Stranded (end-state)",4.5,"= 15 - bypass 5.7 - Hormuz 4.8; reversible escort transits discounted")]
for i,(q,v,n) in enumerate(tp,start=2):
    ws.cell(i,1,q).alignment=LEFT; ws.cell(i,2,v).alignment=CEN; ws.cell(i,3,n).alignment=LEFT
    for c in range(1,4): ws.cell(i,c).border=BORDER
ws.freeze_panes="A2"; autosize(ws,[34,8,52])

# ============================================================ TRAJECTORY ====
ws = wb.create_sheet("Trajectory")
cols=["Week of","OECD stock (mb)","US SPR (mb)","Marker"]; ws.append(cols); style_header(ws,1,4)
tj=[("27 Feb",4100,411,"Eve of war"),("13 Mar",4064,411,"IEA 400 mb release decided"),
    ("03 Apr",4000,400.9,"Global Apr draw -74 mb"),("08 May",3825,384.1,"Record weekly SPR draw 8.6 mb"),
    ("15 May",3787,374.2,"All-time record weekly draw 9.92 mb"),("29 May",3711,357.1,"OECD govt stocks lowest since Dec 1990"),
    ("19 Jun",3598,331.2,"Lowest since 1983; ~46% of action drawn"),("26 Jun",3560,331.2,"As-of; Brent $74.43; no build week yet")]
for i,(w,o,s,m) in enumerate(tj,start=2):
    ws.cell(i,1,w).alignment=CEN; ws.cell(i,2,o).alignment=CEN; ws.cell(i,3,s).alignment=CEN; ws.cell(i,4,m).alignment=LEFT
    for c in range(1,5): ws.cell(i,c).border=BORDER
ws.freeze_panes="A2"; autosize(ws,[10,16,13,42])
ws.cell(11,1,"Inflection: NOT yet turned — OECD and US SPR both fall every print through 26 Jun.").font=Font(italic=True,color="9E2B25")
# (Native chart intentionally omitted: openpyxl's chart XML triggers Excel's
#  "recover content" prompt. The trajectory is plotted on the dashboard instead;
#  select A1:C9 and Insert > Line if you want an in-sheet chart.)

# ============================================================ SCENARIOS =====
ws = wb.create_sheet("Scenarios")
cols=["","S1 — Reopen-fast","S2 — Contested-grinding (base)","S3 — Reclose-hard"]
ws.append(cols); style_header(ws,1,4)
rows=[("Trigger","Ceasefire converts to settlement; IRGC lifts closure/escort","Ceasefire holds but doesn't convert; flow ~4.8 reversible","Ceasefire collapses; physical enforcement / mining"),
      ("Reserve trajectory","Draw stops then reverses; SPR slow refill","Slow continued draw; SPR ~1.29->0.72 mb/d; rho ~0.20","Draw re-accelerates; Gulf re-strands; rho still <<1"),
      ("Brent path","$60-68","$70-85 range, headline spikes decay","Re-rate toward/above $120"),
      ("Restock","Multi-quarter, bypass-gated; SPR refill ~3-4 yr","No restock — net draw continues","Indefinitely deferred; none in 2026"),
      ("Reserve hard clock","n/a — draw halts","~150 mb operational floor ~Nov 26 at acute rate","US SPR->150 mb ~5 Nov 26; OECD authorization ~11 Nov 26")]
for i,row in enumerate(rows,start=2):
    for cidx,v in enumerate(row,1):
        cell=ws.cell(i,cidx,v); cell.alignment=(LEFT if cidx==1 else CEN); cell.border=BORDER
        if cidx==1: cell.font=Font(bold=True)
ws.freeze_panes="B2"; autosize(ws,[18,34,38,38])

# ============================================================ FOOTNOTES =====
ws = wb.create_sheet("Footnotes")
fns=["FOOTNOTES & LIMITATIONS",
 "",
 "F-rho: rho = producer build / consumer draw = 120/615 = 0.195. Verdict insensitive: rho<0.3 across plausible range.",
 "F-at-risk: Eff. at-risk cover = days x (1 - Hormuz-dep) is a WORST-CASE zero-replacement BOUND, not a coverage estimate.",
 "   Japan/Korea 20.5 and Philippines 2.3 are scare-bounds; realistic effective cover is materially higher for diversifiers.",
 "F-netexporter: USA & Netherlands carry '-' on IEA net-import basis (denominator <=0). USA alt: SPR-only ~63d, crude ~129d, total ~259d vs gross imports.",
 "F-Vietnam: published 40.3d (D-basis) diverges >15% from model (~16d). Both shown; headline=higher tier; not averaged.",
 "F-Thailand/Indonesia: Thailand 117 bundles strategic+commercial+in-transit+contracted; hard on-ground ~56d. Indonesia capacity-capped ~25d.",
 "F-Qatar: producer block downgraded Inferred->Modelled-EST (self-disclosed tankage + Wikipedia spec overreached the tag).",
 "F-ledger: original decomposition double-counted (472+615+410 overshot gap by 25%). Corrected to single MECE identity; residual -177 mb (~15%).",
 "F-foregone: flat 9.2 mb/d x118d = 1,080 mb is indefensible vs flow data; time-weighted ~700-850 mb. REFUTED survives.",
 "F-verification: NOT fully independent. IEA OMR / EIA weekly hosts were egress-proxy-403-blocked; anchors are WIRE-ATTRIBUTED (one tier below Verified).",
 "   Re-pull iea.org / eia.gov directly before external use. Self-citation guard held (hormuzstraitmonitor.com / straits.live excluded).",
 "F-redteam: isolated red team raised 10 substantive attacks; all addressed in-run. Verify verdict: ship.",
 "",
 "Provenance: workflow run wf_9642a98b-45e, 52 agents, ~1.9M tokens, as-of 26 Jun 2026.",
]
for i,t in enumerate(fns,1):
    ws.cell(i,1,t).font=Font(bold=(i==1 or "verification" in t.lower()), size=(12 if i==1 else 10),
                             color=("9E2B25" if ("NOT fully independent" in t or i==1) else "1B2B34"))
autosize(ws,[150])

wb.save("OUTPUT_hormuz2026_reserve_intelligence.xlsx")
print("wrote OUTPUT_hormuz2026_reserve_intelligence.xlsx with", len(wb.sheetnames), "sheets:", wb.sheetnames)
